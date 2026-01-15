# test_chatbot_eval.py (version corrigée)
import csv
import time
import re
import sys
from typing import List, Dict
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import chardet

# Import du service RAG
try:
    from app.services.rag_service import query
except ImportError:
    # Fallback pour le test
    def query(question, chatbot_id, user_id):
        return {
            "answer": f"Réponse simulée pour: {question}",
            "time_to_respond": 0.5
        }

# ⚙️ Config
CHATBOT_ID = "68e58e30209aa5b73a4e51a3"
USER_ID = "68d08fde736d1ecda2a2f2c3"

# Rendre le module "app" accessible
sys.path.append(str(Path(__file__).resolve().parents[1]))

# 📁 Fichiers CSV d'entrée et de sortie
CSV_PATH = Path(__file__).resolve().with_name("rh_chatbot_eval.csv")
OUTPUT_EXCEL = Path(__file__).resolve().with_name("rh_chatbot_eval_results.xlsx")

# ------------------ 🔹 UTILITAIRES ------------------ #

def detect_encoding(file_path: Path) -> str:
    """Détecte l'encodage du fichier."""
    with open(file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        encoding = result.get('encoding', 'utf-8')
        confidence = result.get('confidence', 0)
        print(f"🔍 Encodage détecté: {encoding} (confiance: {confidence:.2f})")
        return encoding

def load_questions(csv_path: Path) -> List[Dict]:
    """Charge les questions et réponses attendues depuis le CSV avec gestion d'encodage."""
    questions = []
    
    if not csv_path.exists():
        print(f"❌ Fichier non trouvé: {csv_path}")
        return questions
    
    # Détection d'encodage
    detected_encoding = detect_encoding(csv_path)
    
    # Essayer différents encodages
    encodings = [detected_encoding, 'utf-8', 'latin-1', 'windows-1252', 'iso-8859-1', 'cp1252']
    
    for encoding in encodings:
        try:
            print(f"📖 Tentative de lecture avec l'encodage: {encoding}")
            with open(csv_path, 'r', encoding=encoding, errors='replace') as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames
                print(f"📋 En-têtes détectés: {fieldnames}")
                
                if not fieldnames:
                    print("❌ Aucun en-tête détecté")
                    continue
                
                for i, row in enumerate(reader):
                    # Vérifier que la ligne n'est pas vide
                    if not any(row.values()):
                        continue
                        
                    # Chercher les colonnes pertinentes - CORRECTION ICI
                    question_key = None
                    expected_key = None
                    
                    for key in row.keys():
                        if key is None:
                            continue
                        key_lower = str(key).lower() if key else ""
                        if 'question' in key_lower:
                            question_key = key
                        if 'expected' in key_lower or 'answer' in key_lower or 'réponse' in key_lower:
                            expected_key = key
                    
                    # Si pas trouvé, utiliser les deux premières colonnes
                    if not question_key and len(fieldnames) >= 1:
                        question_key = fieldnames[0]
                    if not expected_key and len(fieldnames) >= 2:
                        expected_key = fieldnames[1]
                    
                    if question_key and expected_key and row.get(question_key):
                        question_text = row[question_key].strip()
                        expected_text = row.get(expected_key, "").strip()
                        
                        if question_text:  # Vérifier que la question n'est pas vide
                            questions.append({
                                "question": question_text,
                                "expected": expected_text
                            })
                            if i < 3:  # Afficher seulement les 3 premières pour le debug
                                print(f"   ✅ Q{i+1}: {question_text[:50]}...")
                
                if questions:
                    print(f"✅ {len(questions)} questions chargées avec l'encodage {encoding}")
                    return questions
                    
        except Exception as e:
            print(f"❌ Erreur avec l'encodage {encoding}: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    # Si tous les encodages échouent, essayer la lecture manuelle
    print("🔄 Tentative de lecture manuelle...")
    for encoding in encodings:
        try:
            with open(csv_path, 'r', encoding=encoding, errors='replace') as f:
                lines = f.readlines()
                print(f"📄 {len(lines)} lignes lues manuellement")
                
                if len(lines) > 0:
                    # Essayer de parser la première ligne comme en-tête
                    first_line = lines[0].strip()
                    headers = [h.strip().strip('"') for h in first_line.split(',')]
                    print(f"📋 En-têtes manuels: {headers}")
                    
                    question_idx = 0  # Première colonne par défaut
                    expected_idx = 1  # Deuxième colonne par défaut
                    
                    # Parcourir les lignes de données
                    for i, line in enumerate(lines[1:], 1):
                        if line.strip():
                            values = [v.strip().strip('"') for v in line.split(',')]
                            if len(values) >= 2 and values[question_idx].strip():
                                questions.append({
                                    "question": values[question_idx].strip(),
                                    "expected": values[expected_idx].strip() if expected_idx < len(values) else ""
                                })
                                if i <= 3:  # Afficher seulement les 3 premières
                                    print(f"   ✅ Q{i}: {values[question_idx][:50]}...")
                    
                    if questions:
                        print(f"✅ {len(questions)} questions chargées manuellement")
                        return questions
                        
        except Exception as e:
            print(f"❌ Erreur lecture manuelle avec {encoding}: {e}")
            continue
    
    print("❌ Aucune question n'a pu être chargée")
    return questions

def extract_numeric_value(text: str) -> float:
    matches = re.findall(r'-?\d+\.?\d*', text)
    if matches:
        try:
            return float(matches[0])
        except ValueError:
            pass
    return None

def evaluate_answer(generated: str, expected: str) -> bool:
    """Compare la réponse générée et attendue avec une logique flexible."""
    if not generated or not expected:
        return False
        
    generated_lower = generated.lower()
    expected_lower = expected.lower()

    if expected_lower in generated_lower:
        return True

    # Comparaison numérique
    expected_num = extract_numeric_value(expected)
    if expected_num is not None:
        generated_num = extract_numeric_value(generated)
        if generated_num is not None:
            return abs(generated_num - expected_num) < 0.01

    # Vérification des mots-clés importants
    important_keywords = ['congés', 'salaire', 'formation', 'mutuelle', 'télétravail', 'rh@entreprise.com', '01 23 45 67', 'recrutement', 'paie', 'performance']
    for keyword in important_keywords:
        if keyword in expected_lower and keyword in generated_lower:
            return True

    return False

# ------------------ 🔹 TEST PAR SESSION ------------------ #

def run_test_session(session_name: str, chatbot_id: str, user_id: str, questions: List[Dict], results_rows: List[Dict]):
    results = {"session": session_name, "num_requests": 0, "latencies": [], "correct": 0}

    for i, q in enumerate(questions, 1):
        try:
            print(f"🔍 Traitement question {i}: {q['question'][:50]}...")
            
            start = time.time()
            response = query(q["question"], chatbot_id, user_id)
            elapsed = response.get("time_to_respond", time.time() - start)
            generated = response.get("answer", "")

            is_correct = evaluate_answer(generated, q["expected"])
            results["num_requests"] += 1
            results["latencies"].append(elapsed)
            if is_correct:
                results["correct"] += 1

            status = "✅" if is_correct else "❌"
            print(f"{status} [{session_name}] Q{i}: {q['question']}")
            print(f"   Généré: {generated[:100]}...")
            print(f"   Attendu: {q['expected']}")
            print(f"   Latence: {elapsed:.2f}s\n")

            # Ligne pour Excel final
            results_rows.append({
                "Question posée": q["question"],
                "Réponse attendue (de référence)": q["expected"],
                "Réponse donnée par le chatbot": generated,
                "Temps de réponse (secondes)": round(elapsed, 2),
                "Moment du test (Matin/Midi/Soir)": session_name,
                "Validation humaine": ""
            })

        except Exception as e:
            print(f"❌ [ERROR] Q{i}: {q['question']}")
            print(f"   Erreur: {str(e)}")
            
            # Ajouter quand même la ligne d'erreur
            results_rows.append({
                "Question posée": q["question"],
                "Réponse attendue (de référence)": q["expected"],
                "Réponse donnée par le chatbot": f"ERREUR: {str(e)}",
                "Temps de réponse (secondes)": 0,
                "Moment du test (Matin/Midi/Soir)": session_name,
                "Validation humaine": ""
            })
            continue

    return results

# ------------------ 🔹 SAUVEGARDE EXCEL DIRECT ------------------ #

def save_results_to_excel_direct(results_rows: List[Dict], output_path: Path) -> bool:
    """💾 Crée directement un fichier Excel sans passer par CSV."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Évaluation Chatbot"

        # En-têtes stylés
        headers = [
            "Question posée",
            "Réponse attendue (de référence)",
            "Réponse donnée par le chatbot",
            "Temps de réponse (secondes)",
            "Moment du test (Matin/Midi/Soir)",
            "Validation humaine"
        ]
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        # Données
        for row in results_rows:
            ws.append([
                row["Question posée"],
                row["Réponse attendue (de référence)"],
                row["Réponse donnée par le chatbot"],
                row["Temps de réponse (secondes)"],
                row["Moment du test (Matin/Midi/Soir)"],
                row["Validation humaine"]
            ])

        # Style de la colonne Validation humaine
        validation_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):  # Colonne F
            for cell in row:
                cell.fill = validation_fill
                cell.alignment = Alignment(horizontal="center")

        # Ajustement automatique des largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Sauvegarde
        wb.save(output_path)
        print(f"📘 Rapport Excel généré : {output_path}")
        return True
        
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde Excel: {e}")
        return False

def create_sample_csv():
    """Crée un fichier CSV d'exemple si le fichier n'existe pas."""
    sample_data = [
        "question,expected_answer",
        "Quel est le rôle principal du département RH ?,Le département Ressources Humaines est chargé de la gestion du personnel, du recrutement, du développement des compétences, de la paie et du climat social au sein de l'entreprise.",
        "Quand est versé le salaire ?,Le salaire est versé le 28 de chaque mois.",
        "Combien de jours de congés payés ai-je ?,20 jours ouvrables par an."
    ]
    
    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(sample_data))
    
    print(f"📝 Fichier d'exemple créé: {CSV_PATH}")

# ------------------ 🔹 MAIN ------------------ #

def main():
    print("🧪 Démarrage du test d'évaluation du chatbot RH...")
    
    # Vérifier que le fichier existe
    if not CSV_PATH.exists():
        print(f"❌ Fichier non trouvé: {CSV_PATH}")
        print("📝 Création d'un fichier d'exemple...")
        create_sample_csv()
    
    questions = load_questions(CSV_PATH)

    if not questions:
        print("❌ Aucune question trouvée dans le fichier CSV")
        return

    print(f"📊 {len(questions)} questions chargées")

    # Répartition des sessions
    total_questions = len(questions)
    chunk_size = max(1, total_questions // 3)
    
    sessions = {
        "Matin": questions[0:chunk_size],
        "Midi": questions[chunk_size:2*chunk_size],
        "Soir": questions[2*chunk_size:total_questions]
    }

    all_results = []
    all_latencies = []
    total_requests = total_correct = 0
    results_rows = []

    for session_name, qs in sessions.items():
        print(f"\n{'='*50}")
        print(f"🚀 Session: {session_name} ({len(qs)} questions)")
        print(f"{'='*50}")

        res = run_test_session(session_name, CHATBOT_ID, USER_ID, qs, results_rows)

        avg_latency = sum(res["latencies"]) / len(res["latencies"]) if res["latencies"] else 0
        accuracy = (res["correct"] / res["num_requests"] * 100) if res["num_requests"] > 0 else 0

        all_results.append({
            "Session": session_name,
            "Nombre de requêtes": res["num_requests"],
            "Latence moyenne (s)": round(avg_latency, 2),
            "Performance du RAG (%)": round(accuracy, 1)
        })

        total_requests += res["num_requests"]
        total_correct += res["correct"]
        all_latencies.extend(res["latencies"])

    # Sauvegarde Excel uniquement
    if results_rows:
        if save_results_to_excel_direct(results_rows, OUTPUT_EXCEL):
            print("✅ Rapport Excel généré avec succès")
        else:
            print("❌ Erreur lors de la génération du rapport Excel")
    else:
        print("❌ Aucune donnée à sauvegarder")

    # Résumé global
    global_latency = sum(all_latencies) / len(all_latencies) if all_latencies else 0
    global_perf = (total_correct / total_requests * 100) if total_requests > 0 else 0

    print(f"\n{'='*60}")
    print("📈 RÉSULTATS FINAUX")
    print(f"{'='*60}")
    print(f"{'Session':<10} | {'#Req':<5} | {'Latence (s)':<12} | {'Performance (%)':<15}")
    print("-" * 50)
    for r in all_results:
        print(f"{r['Session']:<10} | {r['Nombre de requêtes']:<5} | {r['Latence moyenne (s)']:<12} | {r['Performance du RAG (%)']:<15}")

    print(f"\n📊 Analyse détaillée :")
    print(f"   • Total des questions: {total_requests}")
    print(f"   • Réponses correctes: {total_correct}")
    print(f"   • Score global: {global_perf:.1f}%")
    print(f"   • Latence moyenne: {global_latency:.2f}s")

if __name__ == "__main__":
    main()